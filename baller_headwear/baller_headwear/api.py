import json
import frappe
from frappe.utils import get_datetime
from frappe import _, msgprint
from datetime import datetime, date 
from frappe.model.mapper import get_mapped_doc
from frappe.query_builder.functions import Sum
from frappe.utils import cint, cstr, flt, get_link_to_form, getdate, new_line_sep, nowdate
from erpnext.manufacturing.doctype.job_card.job_card import make_material_request
from erpnext.buying.utils import check_on_hold_or_closed_status, validate_for_items
from erpnext.controllers.buying_controller import BuyingController
from erpnext.manufacturing.doctype.work_order.work_order import get_item_details
from erpnext.stock.doctype.item.item import get_item_defaults
from erpnext.stock.stock_balance import get_indented_qty, update_bin_qty
import sys
from frappe.utils import now_datetime
from datetime import datetime, timedelta
from frappe.utils import format_time
from openpyxl import load_workbook
import os
import re

@frappe.whitelist()
def make_stock_entry(source_name, target_doc=None):
	def update_item(obj, target, source_parent):
		qty = (
			flt(flt(obj.stock_qty) - flt(obj.ordered_qty)) / target.conversion_factor
			if flt(obj.stock_qty) > flt(obj.ordered_qty)
			else 0
		)
		target.qty = qty
		target.transfer_qty = qty * obj.conversion_factor
		target.conversion_factor = obj.conversion_factor

		if (
			source_parent.material_request_type == "Material Transfer"
			or source_parent.material_request_type == "Customer Provided"
		):
			target.t_warehouse = obj.warehouse
		else:
			target.s_warehouse = obj.warehouse

		if source_parent.material_request_type == "Customer Provided":
			target.allow_zero_valuation_rate = 1

		if source_parent.material_request_type == "Material Transfer":
			target.s_warehouse = obj.from_warehouse

	def set_missing_values(source, target):
		# target.purpose = source.material_request_type
		target.purpose = "Material Transfer for Manufacture"
		target.from_warehouse = source.set_from_warehouse
		target.to_warehouse = source.set_warehouse

		if source.job_card:
			target.purpose = "Material Transfer for Manufacture"

		if source.material_request_type == "Customer Provided":
			target.purpose = "Material Receipt"

		target.set_transfer_qty()
		target.set_actual_qty()
		target.calculate_rate_and_amount(raise_error_if_no_rate=False)
		target.stock_entry_type = target.purpose
		target.set_job_card_data()

		if source.job_card:
			job_card_details = frappe.get_all(
				"Job Card", filters={"name": source.job_card}, fields=["bom_no", "for_quantity"]
			)

			if job_card_details and job_card_details[0]:
				target.bom_no = job_card_details[0].bom_no
				target.fg_completed_qty = job_card_details[0].for_quantity
				target.from_bom = 1

	doclist = get_mapped_doc(
		"Material Request",
		source_name,
		{
			"Material Request": {
				"doctype": "Stock Entry",
				"validation": {
					"docstatus": ["=", 1],
					"material_request_type": [
						"in",
						["Material Transfer", "Material Issue", "Customer Provided"],
					],
				},
			},
			"Material Request Item": {
				"doctype": "Stock Entry Detail",
				"field_map": {
					"name": "material_request_item",
					"parent": "material_request",
					"uom": "stock_uom",
					"job_card_item": "job_card_item",
				},
				"postprocess": update_item,
				"condition": lambda doc: (
					flt(doc.ordered_qty, doc.precision("ordered_qty"))
					< flt(doc.stock_qty, doc.precision("ordered_qty"))
				),
			},
		},
		target_doc,
		set_missing_values,
	)

	return doclist

# @frappe.whitelist()
# def set_custom_id_fields_for_transaction_date(doc, method):

#     if doc.transaction_date:

#         transaction_date = datetime.strptime(doc.transaction_date, '%Y-%m-%d')

#         month = transaction_date.month
#         year = transaction_date.year
        

#         year_formatted = str(year)[-2:]  
#         month_formatted = f"{month:02d}" 
        

#         doc.custom_id_month = month_formatted
#         doc.custom_id_year = year_formatted

@frappe.whitelist()
def set_custom_id_fields_for_transaction_date(doc, method):
    if doc.transaction_date:
        if isinstance(doc.transaction_date, str):
            transaction_date = datetime.strptime(doc.transaction_date, '%Y-%m-%d')
        else:
            transaction_date = doc.transaction_date

        month = transaction_date.month
        year = transaction_date.year

        year_formatted = str(year)[-2:]  
        month_formatted = f"{month:02d}" 

        doc.custom_id_month = month_formatted
        doc.custom_id_year = year_formatted
        naming_series = doc.naming_series
        if doc.naming_series:
            doc.naming_series = naming_series.replace("YYYY", str(year))

# @frappe.whitelist()
# def set_custom_id_fields_for_posting_date(doc, method):

#     if doc.posting_date:

#         posting_date = datetime.strptime(doc.posting_date, '%Y-%m-%d')

#         month = posting_date.month
#         year = posting_date.year
        

#         year_formatted = str(year)[-2:]  
#         month_formatted = f"{month:02d}" 
        

#         doc.custom_id_month = month_formatted
#         doc.custom_id_year = year_formatted

@frappe.whitelist()
def set_custom_id_fields_for_posting_date(doc, method):
    if doc.posting_date:
        if isinstance(doc.posting_date, str):
            posting_date = datetime.strptime(doc.posting_date, '%Y-%m-%d')
        else:
            posting_date = doc.posting_date

        month = posting_date.month
        year = posting_date.year
        year_formatted = str(year)[-2:]  
        month_formatted = f"{month:02d}"
        doc.custom_id_month = month_formatted
        doc.custom_id_year = year_formatted

        naming_series = doc.naming_series
        if doc.naming_series:
            doc.naming_series = naming_series.replace("YYYY", str(year))
             


def parse_date(date_input):
    if date_input is None:
        return None

    if isinstance(date_input, datetime):
        return date_input

    if isinstance(date_input, date):
        return datetime.combine(date_input, datetime.min.time())

    if isinstance(date_input, (int, float)):
        # Treat numeric input as a timestamp (in seconds)
        return datetime.fromtimestamp(float(date_input))

    # Convert anything else to string and try parsing
    date_str = str(date_input).strip()
    formats = [
        '%Y-%m-%d %H:%M:%S.%f',
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d',
        '%Y/%m/%d',
    ]
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue

    raise ValueError(f"Could not parse date: {date_input!r}")



@frappe.whitelist() 
def set_custom_id_fields_for_work_order(doc, method):
    if doc.planned_start_date:
        planned_start_date = parse_date(doc.planned_start_date)

        month = planned_start_date.month
        year = planned_start_date.year

        year_formatted = str(year)[-2:]  
        month_formatted = f"{month:02d}" 

        doc.custom_id_month = month_formatted
        doc.custom_id_year = year_formatted


@frappe.whitelist()
def set_custom_id_fields_for_posting_date_jv(doc, method):
    company_settings = frappe.get_doc("Company Settings")
    if doc.posting_date:
        if isinstance(doc.posting_date, str):
            posting_date = datetime.strptime(doc.posting_date, '%Y-%m-%d').date()
        elif isinstance(doc.posting_date, datetime):  
            posting_date = doc.posting_date.date()
        elif isinstance(doc.posting_date, date):
            posting_date = doc.posting_date
        else:
            raise ValueError("Posting date is neither a string nor a datetime.date object.")

        posting_date = datetime.combine(posting_date, datetime.min.time())

        month = posting_date.month
        year = posting_date.year

        year_formatted = str(year)[-2:]  
        month_formatted = f"{month:02d}" 

        doc.custom_id_month = month_formatted
        doc.custom_id_year = year_formatted
        
        if doc.naming_series == "ACC-JV-2024-":
            doc.naming_series = "ACC-JV-2024-" 
        elif doc.company == company_settings.vietnam_company_name:
            doc.naming_series = ".{custom_abbr}.-JV-.{custom_id_year}.-.{custom_id_month}.-.####"
        else:
            doc.naming_series = "ACC-JV-" + str(year) + '-'
            
@frappe.whitelist()
def set_custom_id_fields_for_posting_date_payment_entry(doc, method):
    company_settings = frappe.get_doc("Company Settings")
    if doc.posting_date:
        posting_date = doc.posting_date
        if isinstance(doc.posting_date, str):
            posting_date = datetime.strptime(doc.posting_date, '%Y-%m-%d')
        month = posting_date.month
        year = posting_date.year
        year_formatted = str(year)[-2:]  
        month_formatted = f"{month:02d}" 
        

        doc.custom_id_month = month_formatted
        doc.custom_id_year = year_formatted

        if doc.naming_series == "ACC-PAY-2024-":
            doc.naming_series = "ACC-PAY-2024-" 
        elif doc.company == company_settings.vietnam_company_name:
            doc.naming_series = ".{custom_abbr}.-PAY-.{custom_id_year}.-.{custom_id_month}.-.####"
        else:
            doc.naming_series = "ACC-PAY-" + str(year) + '-'

@frappe.whitelist()
def set_custom_id_fields_for_asset(doc, method):
    company_settings = frappe.get_doc("Company Settings")

    if doc.purchase_date:
        purchase_date = doc.purchase_date
        if isinstance(purchase_date, date):
            # It's already a date object, keep as is
            pass
        else:
            # Convert from string (if it's a string or something else)
            purchase_date = datetime.strptime(str(purchase_date), '%Y-%m-%d').date()


        month = purchase_date.month
        year = purchase_date.year
        

        year_formatted = str(year)[-2:]  
        month_formatted = f"{month:02d}" 
        

        doc.custom_id_month = month_formatted
        doc.custom_id_year = year_formatted

        if doc.naming_series == "ACC-ASS-2024-":
            doc.naming_series = "ACC-ASS-2024-" 
        elif doc.company == company_settings.vietnam_company_name:
            doc.naming_series = ".{custom_abbr}.-ASS-.{custom_id_year}.-.{custom_id_month}.-.####"
        else:
            doc.naming_series = "ACC-ASS-" + str(year) + '-'


@frappe.whitelist()
def get_item_image(item_code):
    image = frappe.db.get_value("Item", item_code, "image")
    return image

@frappe.whitelist()
def get_work_orders_for_cutting(from_date=None, to_date=None, item=None):
    filters = {
        "status": "Not Started"
    }

    if item:
        filters["production_item"] = item

    if from_date and to_date:
        from_datetime = get_datetime(from_date + " 00:00:00")
        to_datetime = get_datetime(to_date + " 23:59:59")
        filters["planned_start_date"] = ["between", [from_datetime, to_datetime]]

    if not item and not (from_date and to_date):
        return []

    work_orders = frappe.get_list(
        "Work Order",
        filters=filters,
        fields=["name"],
        order_by="planned_start_date asc"
    )

    result = []

    for wo in work_orders:
        work_order_doc = frappe.get_doc("Work Order", wo.name)
        for item in work_order_doc.required_items:
            item_doc = frappe.get_doc("Item", item.item_code)
            if item_doc.item_group == "Fabric":
                result.append({
                    "work_order": wo.name,
                    "item_fabric": item.item_code,
                    "qty": item.required_qty
                })

    return result

# @frappe.whitelist()
# def get_work_orders_for_cutting(from_date, to_date):
#     from_datetime = get_datetime(from_date + " 00:00:00")
#     to_datetime = get_datetime(to_date + " 23:59:59")

#     work_orders = frappe.get_list(
#         "Work Order",
#         filters={
#             "planned_start_date": ["between", [from_datetime, to_datetime]],
#             "status": "Not Started"
#         },
#         fields=["name"],
#         order_by="planned_start_date asc"
#     )

#     result = []

#     for wo in work_orders:
#         work_order_doc = frappe.get_doc("Work Order", wo.name)
#         for item in work_order_doc.required_items:
#             item_doc = frappe.get_doc("Item", item.item_code)
#             if item_doc.item_group == "Fabric":
#                 result.append({
#                     "work_order": wo.name,
#                     "item_fabric": item.item_code,
#                     "qty": item.required_qty
#                 })

#     return result


def fmt_date(d):
    return d.strftime("%Y-%m-%d") if isinstance(d, datetime) else None

def fmt_time(t):
    new_time = t
    if isinstance(t, datetime):
        new_time = t.strftime("%H:%M:%S")
    if isinstance(t, str):
        new_time = datetime.strptime(t, "%I:%M:%S %p").strftime("%H:%M:%S")
    return new_time


def process_bulk(item_list, item_list_new, item_map):
    from erpnext.manufacturing.doctype.work_order.work_order import make_stock_entry
    from erpnext.stock.doctype.stock_entry.stock_entry import get_warehouse_details
    for item in item_list:
        item_code = item.get('product_code')
        work_orders = frappe.get_all(
            "Work Order",
            filters={
                "production_item": ["like", f"%{item_code}%"],
                "status": ["!=", "Completed"],
            },
            fields= [
                "name",
                "status",
                "bom_no",
                "qty",
                "production_item",
                "material_transferred_for_manufacturing",
                "produced_qty",
                "planned_start_date",
                "planned_end_date"
            ]
        )

        priority_wos = []
        normal_wos = []
        for wo in work_orders:
            item_group = frappe.db.get_value(
                "Item",
                wo['production_item'],
                "item_group"
            )

            if item_group != 'Semi-finished':
                continue

            count = frappe.db.sql("""
                SELECT COUNT(woi.name)
                FROM `tabWork Order Item` woi
                INNER JOIN `tabItem` i ON i.name = woi.item_code
                WHERE woi.parent = %s
            """, (wo.name))[0][0]

            count_diff_semi = frappe.db.sql(
                """
                SELECT COUNT(woi.name)
                FROM `tabWork Order Item` woi
                INNER JOIN `tabItem` i ON i.name = woi.item_code
                WHERE woi.parent = %s
                AND i.item_group != %s
                """,
                (wo.name, "Semi-finished")
            )[0][0]

            if count_diff_semi > 0:
                wo.required_items_count = count
                priority_wos.append(wo)
            elif count > 1:
                priority_wos.append(wo)
            else:
                normal_wos.append(wo)

        for wo in priority_wos:
            wo_in_new_list = item_map.get(wo.name)
            if not wo_in_new_list:
                continue
            try:
                resp = make_stock_entry(wo.name, 'Material Transfer for Manufacture', 0)
                stock_entry_items = resp['items']
                if not stock_entry_items:
                    continue
                
                for stock_item in stock_entry_items:
                    item_group = frappe.db.get_value(
                        "Item",
                        stock_item['item_code'],
                        "item_group"
                    )
                    stock_item.s_warehouse = 'Main Store - BHV'
                    if stock_item['item_code'] == 'Piping 3.0-9672-135':
                        stock_item.item_code = 'Piping 2.9-9672-135'

                    if item_group == 'Semi-finished':
                        stock_item.s_warehouse = 'Semi-finished Goods  - BHV'
                        
                        args = {
                                "item_code":stock_item['item_code'],
                                "warehouse":"Semi-finished Goods  - BHV",
                                "transfer_qty":stock_item['qty'],
                                "serial_and_batch_bundle":None,
                                "qty":(0-stock_item['qty']),
                                "posting_date":wo_in_new_list.get('posting_date'),
                                "posting_time":wo_in_new_list.get('posting_time'),
                                "company":"Baller Headwear Vietnam Ltd",
                                "voucher_type":"Stock Entry",
                                "voucher_no":"new-stock-entry-detail-mpfrecxxiz",
                                "allow_zero_valuation":1
                            }
                        result = get_warehouse_details(args)
                        stock_item.basic_rate = result.get("basic_rate")
                
                doc = frappe.get_doc(resp)  
                doc.set_posting_time = 1
                doc.fg_completed_qty = 0
                doc.posting_date = wo_in_new_list.get('posting_date')
                doc.posting_time = wo_in_new_list.get('posting_time')
                doc.insert(ignore_permissions=True)
                doc.submit()
                frappe.db.commit()
            except frappe.ValidationError:
                frappe.db.rollback()
                continue
            except Exception as e:
                frappe.db.rollback()
                continue

        for wo in priority_wos:
            wo_in_new_list = item_map.get(wo.name)
            if not wo_in_new_list:
                continue
            try:
                resp = make_stock_entry(wo.name, 'Manufacture', wo.qty)
                doc = frappe.get_doc(resp)  
                doc.set_posting_time = 1
                doc.posting_date = wo_in_new_list.get('posting_date')
                doc.posting_time = wo_in_new_list.get('posting_time')
                doc.insert(ignore_permissions=True)
                doc.submit()
                frappe.db.commit()
            except frappe.ValidationError:
                frappe.db.rollback()
                print("ERROR Manufacture",  wo.name)
                print("WO", wo.name, 'posting_date', wo_in_new_list.get('posting_date'), 'posting_time', wo_in_new_list.get('posting_time'))
                continue
            except Exception as e:
                frappe.db.rollback()
                print("WO", wo.name, 'posting_date', item.get('posting_date'), 'posting_time', item.get('posting_time'))
                continue

        for wo in normal_wos:
            wo_in_new_list = item_map.get(wo.name)
            if not wo_in_new_list:
                continue
            try:
                resp = make_stock_entry(wo.name, 'Material Transfer for Manufacture', 0)
                stock_entry_items = resp['items']
                if not stock_entry_items:
                    continue
                for stock_item in stock_entry_items:
                    item_group = frappe.db.get_value(
                        "Item",
                        stock_item['item_code'],
                        "item_group"
                    )
                    stock_item.s_warehouse = 'Main Store - BHV'
                    if stock_item['item_code'] == 'Piping 3.0-9672-135':
                        stock_item.item_code = 'Piping 2.9-9672-135'

                    if item_group == 'Semi-finished':
                        stock_item.s_warehouse = 'Semi-finished Goods  - BHV'
                        
                        args = {
                                "item_code":stock_item['item_code'],
                                "warehouse":"Semi-finished Goods  - BHV",
                                "transfer_qty":stock_item['qty'],
                                "serial_and_batch_bundle":None,
                                "qty":(0-stock_item['qty']),
                                "posting_date":wo_in_new_list.get('posting_date'),
                                "posting_time":wo_in_new_list.get('posting_time'),
                                "company":"Baller Headwear Vietnam Ltd",
                                "voucher_type":"Stock Entry",
                                "voucher_no":"new-stock-entry-detail-mpfrecxxiz",
                                "allow_zero_valuation":1
                            }
                        result = get_warehouse_details(args)
                        stock_item.basic_rate = result.get("basic_rate")
                
                doc = frappe.get_doc(resp)  
                doc.set_posting_time = 1
                doc.fg_completed_qty = 0
                doc.posting_date = wo_in_new_list.get('posting_date')
                doc.posting_time = wo_in_new_list.get('posting_time')
                doc.insert(ignore_permissions=True)
                doc.submit()
                frappe.db.commit()
            except frappe.ValidationError:
                frappe.db.rollback()
                print("ERROR Material Transfer for Manufacture",  wo.name)
                print("WO", wo.name, 'posting_date', wo_in_new_list.get('posting_date'), 'posting_time', wo_in_new_list.get('posting_time'))
                continue
            except Exception as e:
                frappe.db.rollback()
                print("WO", wo.name, 'posting_date', item.get('posting_date'), 'posting_time', item.get('posting_time'))
                continue

        for wo in normal_wos:
            wo_in_new_list = item_map.get(wo.name)
            if not wo_in_new_list:
                continue
            try:
                resp = make_stock_entry(wo.name, 'Manufacture', wo.qty)
                doc = frappe.get_doc(resp)  
                doc.set_posting_time = 1
                doc.posting_date = wo_in_new_list.get('posting_date')
                doc.posting_time = wo_in_new_list.get('posting_time')
                doc.insert(ignore_permissions=True)
                doc.submit()
                frappe.db.commit()
            except frappe.ValidationError:
                frappe.db.rollback()
                continue
            except Exception as e:
                frappe.db.rollback()
                continue


    for item in item_list:
        item_code = item.get('product_code')
        work_orders = frappe.get_all(
            "Work Order",
            filters={
                "production_item": ["like", f"{item_code}%"],
                "status": ["!=", "Completed"],
            },
            fields= [
                "name",
                "status",
                "bom_no",
                "qty",
                "production_item",
                "material_transferred_for_manufacturing",
                "produced_qty",
                "planned_start_date",
                "planned_end_date"
            ]
        )

        for wo in work_orders:
            wo_in_new_list = item_map.get(wo.name)
            if not wo_in_new_list:
                continue
            try:
                resp = make_stock_entry(wo.name, 'Material Transfer for Manufacture', 0)
                stock_entry_items = resp['items']
                if not stock_entry_items:
                    continue
                for stock_item in stock_entry_items:
                    item_group = frappe.db.get_value(
                        "Item",
                        stock_item['item_code'],
                        "item_group"
                    )
                    stock_item.s_warehouse = 'Main Store - BHV'
                    if stock_item['item_code'] == 'Piping 3.0-9672-135':
                        stock_item.item_code = 'Piping 2.9-9672-135'

                    if item_group == 'Semi-finished':
                        stock_item.s_warehouse = 'Semi-finished Goods  - BHV'
                        
                        args = {
                                "item_code":stock_item['item_code'],
                                "warehouse":"Semi-finished Goods  - BHV",
                                "transfer_qty":stock_item['qty'],
                                "serial_and_batch_bundle":None,
                                "qty":(0-stock_item['qty']),
                                "posting_date":wo_in_new_list.get('posting_date'),
                                "posting_time":wo_in_new_list.get('posting_time'),
                                "company":"Baller Headwear Vietnam Ltd",
                                "voucher_type":"Stock Entry",
                                "voucher_no":"new-stock-entry-detail-mpfrecxxiz",
                                "allow_zero_valuation":1
                            }
                        result = get_warehouse_details(args)
                        stock_item.basic_rate = result.get("basic_rate")
                
                doc = frappe.get_doc(resp)  
                doc.set_posting_time = 1
                doc.fg_completed_qty = 0
                doc.posting_date = wo_in_new_list.get('posting_date')
                doc.posting_time = wo_in_new_list.get('posting_time')
                doc.insert(ignore_permissions=True)
                doc.submit()
                frappe.db.commit()
            except frappe.ValidationError:
                frappe.db.rollback()
                print("ERROR Material Transfer for Manufacture",  wo.name)
                print("WO", wo.name, 'posting_date', wo_in_new_list.get('posting_date'), 'posting_time', wo_in_new_list.get('posting_time'))
                continue
            except Exception as e:
                frappe.db.rollback()
                print("WO", wo.name, 'posting_date', item.get('posting_date'), 'posting_time', item.get('posting_time'))
                continue

        for wo in work_orders:
            wo_in_new_list = item_map.get(wo.name)
            if not wo_in_new_list:
                continue
            try:
                resp = make_stock_entry(wo.name, 'Manufacture', wo.qty)
                doc = frappe.get_doc(resp)  
                doc.set_posting_time = 1
                doc.posting_date = wo_in_new_list.get('posting_date')
                doc.posting_time = wo_in_new_list.get('posting_time')
                doc.insert(ignore_permissions=True)
                doc.submit()
                frappe.db.commit()
            except frappe.ValidationError:
                frappe.db.rollback()
                print("ERROR Manufacture",  wo.name)
                print("WO", wo.name, 'posting_date', wo_in_new_list.get('posting_date'), 'posting_time', wo_in_new_list.get('posting_time'))
                continue
            except Exception as e:
                frappe.db.rollback()
                print("WO", wo.name, 'posting_date', item.get('posting_date'), 'posting_time', item.get('posting_time'))
                continue
    
@frappe.whitelist(allow_guest=True)
def run_job():
    data_not_dup_rows = frappe.db.sql("""
        SELECT DISTINCT product_code
        FROM `tabAuto Complete Wo Jobs`
    """, as_dict=True)

    data_rows = frappe.db.sql("""
        SELECT *
        FROM `tabAuto Complete Wo Jobs`
    """, as_dict=True)

    item_list = data_not_dup_rows
    item_list_new = data_rows
    item_map = {item["wo"]: item for item in item_list_new}
    BATCH_SIZE = 3
    for i in range(0, len(item_list), BATCH_SIZE):
        batch = item_list[i:i + BATCH_SIZE]
        frappe.enqueue(
            "baller_headwear.baller_headwear.api.process_bulk",
            queue="long",
            timeout=3600,
            item_list=batch,
            item_map=item_map,
            item_list_new=item_list_new
        )

def read_file():
    from openpyxl import load_workbook
    import re
    file_path = '/home/son96/works/erp.bellar.15/1313.xlsx'
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb.active   # hoặc wb["Sheet1"]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_raw, item, plan, mfg_raw, qty, completed = row[:6]

        date = date_raw.strftime("%Y-%m-%d") if isinstance(date_raw, datetime) \
            else datetime.strptime(date_raw, "%d/%m/%Y").strftime("%Y-%m-%d")

        production_plans = [
            p.strip()
            for p in re.split(r"[,\n]", str(mfg_raw))
            if p.strip()
        ]

        data.append({
            "date": date,
            "item_code": str(item).strip(),
            "plan": str(plan).strip(),
            "production_plans": production_plans,
            "qty": int(qty or 0),
            "completed_qty": int(completed or 0)
        })

    return data

def process_bulk_melin(new_data):
    from erpnext.manufacturing.doctype.work_order.work_order import make_stock_entry
    from erpnext.stock.doctype.stock_entry.stock_entry import get_warehouse_details
    for item in new_data:
        item_code = item.get('item_code')
        qty = item.get('qty')
        production_plans = item.get('production_plans')
        if not production_plans:
            continue

        for plan in production_plans:
            assembly_items = frappe.db.sql("""
                SELECT
                    item_code,
                    bom_no,
                    planned_qty,
                    warehouse
                FROM `tabProduction Plan Item`
                WHERE parent = %s
                AND item_code = %s
                AND planned_qty = %s
            """, (plan, item_code, qty), as_dict=True)

            if assembly_items:
                work_orders = frappe.get_all(
                    "Work Order",
                    filters={
                        "production_plan": plan,
                        "qty": qty,
                        "production_item": item_code,
                        "status": ["!=", "Completed"],
                    },
                    fields= [
                        "name",
                        "status",
                        "bom_no",
                        "qty",
                        "production_item",
                        "material_transferred_for_manufacturing",
                        "produced_qty",
                        "planned_start_date",
                        "planned_end_date"
                    ],
                    limit=1
                )

                # for wo in work_orders:
                #     try:
                #         resp = make_stock_entry(wo.name, 'Material Transfer for Manufacture', 0)
                #         stock_entry_items = resp['items']
                #         if not stock_entry_items:
                #             continue
                #         for stock_item in stock_entry_items:
                #             item_group = frappe.db.get_value(
                #                 "Item",
                #                 stock_item['item_code'],
                #                 "item_group"
                #             )
                #             stock_item.s_warehouse = 'Main Store - BHV'
                #             if item_group == 'Semi-finished':
                #                 stock_item.s_warehouse = 'Semi-finished Goods  - BHV'
                                
                #                 args = {
                #                         "item_code":stock_item['item_code'],
                #                         "warehouse":"Semi-finished Goods  - BHV",
                #                         "transfer_qty":stock_item['qty'],
                #                         "serial_and_batch_bundle":None,
                #                         "qty":(0-stock_item['qty']),
                #                         "posting_date": item.get('posting_date'),
                #                         "posting_time":'10:32',
                #                         "company":"Baller Headwear Vietnam Ltd",
                #                         "voucher_type":"Stock Entry",
                #                         "voucher_no":"new-stock-entry-detail-mpfrecxxiz",
                #                         "allow_zero_valuation":1
                #                     }
                #                 result = get_warehouse_details(args)
                #                 stock_item.basic_rate = result.get("basic_rate")
                        
                #         doc = frappe.get_doc(resp)  
                #         doc.set_posting_time = 1
                #         doc.fg_completed_qty = 0
                #         doc.posting_date = item.get('date')
                #         doc.insert(ignore_permissions=True)
                #         doc.submit()
                #         frappe.db.commit()
                #     except frappe.ValidationError as e:
                #         frappe.db.rollback()
                #         message=f"WO: {wo.name}\n{frappe.get_traceback()}"
                #         name = item.get('name')

                #         frappe.db.sql("""
                #             UPDATE `tabAuto Melin Wo`
                #             SET error = %s
                #             WHERE name = %s
                #         """, (message, name))

                #         frappe.db.commit()
                #         continue
                #     except Exception as e:
                #         frappe.db.rollback()
                #         continue

                for wo in work_orders:
                    try:
                        resp = make_stock_entry(wo.name, 'Manufacture', item.get('completed_qty'))
                        doc = frappe.get_doc(resp)  
                        doc.set_posting_time = 1
                        doc.posting_date = item.get('date')
                        doc.insert(ignore_permissions=True)
                        doc.submit()
                        frappe.db.commit()
                    except frappe.ValidationError as e:
                        frappe.db.rollback()
                        print("ERROR Manufacture",  wo.name)
                        print(item.get('name'))
                        message=f"WO: {wo.name}\n{frappe.get_traceback()}"
                        name = item.get('name')

                        frappe.db.sql("""
                            UPDATE `tabAuto Melin Wo`
                            SET error = %s
                            WHERE name = %s
                        """, (message, name))

                        frappe.db.commit()
                        continue
                    except Exception as e:
                        frappe.db.rollback()
                        continue


@frappe.whitelist(allow_guest=True)
def run_melin():
    data = frappe.db.sql("""
        SELECT *
        FROM `tabAuto Melin Wo`
    """, as_dict=True)
   
    new_data = []
    for item in data:
        production_plans = [
            p.strip()
            for p in re.split(r"[,\n]", str(item.get('production_plans')))
            if p.strip()
        ]

        new_data.append({
            "name": item.get('name'),
            "date": item.get('date'),
            "item_code": str(item.get('item_code')).strip(),
            "production_plans": production_plans,
            "qty": int(item.get('qty') or 0),
            "completed_qty": int(item.get('completed_qty') or 0)
        })
    
    BATCH_SIZE = 3
    for i in range(0, len(new_data), BATCH_SIZE):
        batch = new_data[i:i + BATCH_SIZE]
        frappe.enqueue(
            "baller_headwear.baller_headwear.api.process_bulk_melin",
            queue="long",
            timeout=3600,
            new_data=batch
        )

@frappe.whitelist()
def create_material_request_with_jobcard(job_cards, workstation, user, required_datetime):
    job_cards = json.loads(job_cards)
    department = frappe.db.get_value("Employee", {"user_id": user}, "department")
    created_mrs = []
    errors = []
    employee_name = frappe.db.get_value(
        "Employee",
        {"user_id": user},
        "employee_name"
    )
    if job_cards:
        for jobcard in job_cards:
            try:
                mr = make_material_request(jobcard)
                if mr:
                    doc = frappe.get_doc(mr)
                    doc.schedule_date = required_datetime
                    doc.custom_regular = 1
                    doc.custom_requester = employee_name
                    doc.custom_requester_dept = department
                    doc.custom_parent_department = ''
                    doc.set_from_warehouse = 'Main Store - BHV'
                    doc.set_warehouse = 'Work In Progress - BHV'
                    doc.insert(ignore_permissions=True)
                    doc.submit()
                    frappe.db.set_value(doc.doctype, doc.name, "workflow_state", "Submited")
                    created_mrs.append(doc.name)

            except Exception as e:
                errors.append(f"{jobcard}: {str(e)}")
                print(frappe.get_traceback())


    return {
        "created": created_mrs,
        "errors": errors
    }

@frappe.whitelist()
def create_material_request_for_transfer_with_jobcard(items, user, required_datetime):
    from frappe.utils import flt
    from collections import defaultdict
    import json
    if isinstance(items, str):
        items = json.loads(items)

    department = frappe.db.get_value("Employee", {"user_id": user}, "department")
    employee_name = frappe.db.get_value("Employee", {"user_id": user}, "employee_name")
    company = frappe.db.get_value("Department", department, "company")
    created_mrs = []
    errors = []

    grouped_items = defaultdict(list)
    for item in items:
        job_card = item.get("job_card")
        if job_card:
            grouped_items[job_card].append(item)

    for job_card, job_items in grouped_items.items():
        try:
            mr_name = make_material_request(job_card)
            mr_name.items = []

            if not mr_name:
                errors.append(f"{job_card}: MR not created")
                continue

            mr = frappe.get_doc(mr_name)
            mr.material_request_type = "Material Transfer"
            mr.schedule_date = required_datetime
            mr.custom_not_regular = 1
            mr.company = company or 'Baller Headwear Vietnam Ltd'
            mr.custom_requester = employee_name
            mr.custom_requester_dept = department
            mr.custom_parent_department = ''
            mr.set_from_warehouse = 'Main Store - BHV'
            mr.set_warehouse = 'Work In Progress - BHV'

            for item in job_items:
                qty = flt(item.get("additional_qty"))
                if qty <= 0:
                    continue

                mr.append("items", {
                    "item_code": item.get("item_code"),
                    "uom": item.get("uom"),
                    "qty": qty,
                    "schedule_date": required_datetime,
                    "from_warehouse": item.get("warehouse") or mr.set_from_warehouse,
                    "warehouse": mr.set_warehouse,
                    "job_card": job_card
                })

            if not mr.items:
                frappe.log_error(f"No valid items for Job Card {job_card}", "MR Warning")
                continue

            mr.insert(ignore_permissions=True)
            mr.submit()
            created_mrs.append(mr.name)

        except Exception as e:
            errors.append(f"{job_card}: {str(e)}")
            frappe.log_error(frappe.get_traceback(), f"MR Error - {job_card}")

    return {"created": created_mrs, "errors": errors}

@frappe.whitelist()
def get_jobcard_raw_materials(job_cards):
    job_cards = json.loads(job_cards)
    items = []
    for jc in job_cards:
        doc = frappe.get_doc("Job Card", jc)
        for row in doc.items:
            items.append({
                "job_card": jc,
                "production_item": doc.production_item,
                "warehouse": row.source_warehouse,
                "item_code": row.item_code,
                "item_name": row.item_name,
                "qty": row.required_qty,
                "uom": row.uom,
                "transferred_qty": row.transferred_qty
            })

    return items